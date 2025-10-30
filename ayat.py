from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardMarkup
from datetime import datetime, timedelta
from fpdf import FPDF
from telegram.ext import (
    ApplicationBuilder, CommandHandler, MessageHandler,
    ConversationHandler, ContextTypes, filters, CallbackQueryHandler
)
import openpyxl
import os
import asyncio
from telegram.constants import ChatAction
import matplotlib.pyplot as plt


# --- Admin setup ---
ADMIN_ID = 1971668737 # Replace with your admin Telegram ID

# --- Conversation states -
NAME, PHONE, CITY, NID, TIN = range(5)
# For prospect adding
PROSPECT_NAME, PROSPECT_PHONE, PROSPECT_INTEREST, PROSPECT_COMMENT = range(100, 104)
BROADCAST = 200
BROADCAST_CHOICE, BROADCAST_MESSAGE = range(200, 202)



# --- Excel setup ---
EXCEL_FILE = "freelancers.xlsx"
PROSPECT_FILE = "prospects.xlsx"

if not os.path.exists(EXCEL_FILE):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Full Name", "Phone Number", "City", "National ID", "TIN Number", "Telegram ID", "Date Registered", "Verified"])
    wb.save(EXCEL_FILE)


if not os.path.exists(PROSPECT_FILE):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Freelancer Telegram ID", "Prospect Name", "Phone Number", "Interest", "Comment", "Date Added"])
    wb.save(PROSPECT_FILE)

# --- Officer Excel files setup ---
OFFICER_FILES = ["officer1.xlsx", "officer2.xlsx", "officer3.xlsx", "officer4.xlsx"]

# --- Hardcoded Officer Info ---
OFFICERS = {
    "officer1.xlsx": {"name": "Alice Desta", "phone": "+251911000111", "username": "@AliceDesta"},
    "officer2.xlsx": {"name": "Biniam Teshome", "phone": "+251911000222", "username": "@BiniamT"},
    "officer3.xlsx": {"name": "Helen Abebe", "phone": "+251911000333", "username": "@HelenA"},
    "officer4.xlsx": {"name": "Samuel Getachew", "phone": "+251911000444", "username": "@SamuelG"},
}

for file in OFFICER_FILES:
    if not os.path.exists(file):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Full Name", "Phone Number", "City", "National ID", "TIN Number", "Telegram ID"])
        wb.save(file)

# --- Admin Report Functions ---

def filter_recent_rows(ws, date_col_index, days):
    """Filter rows from Excel within given number of days."""
    now = datetime.now()
    cutoff = now - timedelta(days=days)
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        date_str = row[date_col_index]
        if not date_str:
            continue
        try:
            date = datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S")
            if date >= cutoff:
                rows.append(row)
        except:
            continue
    return rows

def filter_new_freelancers(ws, days=1):
    now = datetime.now()
    cutoff = now - timedelta(days=days)
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        date_str = row[6]  # 7th column = Date Registered
        if not date_str:
            continue
        try:
            date = datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S")
            if date >= cutoff:
                rows.append(row)
        except:
            continue
    return rows



def create_summary_pdf(daily_count, weekly_count, monthly_count, file_name="summary.pdf"):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", "B", 16)
    pdf.cell(0, 10, "Freelancer & Prospect Summary Report", ln=True, align="C")
    pdf.ln(10)
    pdf.set_font("Arial", "", 12)
    pdf.cell(0, 10, f"Daily (Last 24 hrs): {daily_count}", ln=True)
    pdf.cell(0, 10, f"Weekly (Last 7 days): {weekly_count}", ln=True)
    pdf.cell(0, 10, f"Monthly (Last 30 days): {monthly_count}", ln=True)
    pdf.output(file_name)
    return file_name


def create_trend_pdf(timeframe="weekly", file_name="trend_report.pdf"):
    # Load Excel
    wb_f = openpyxl.load_workbook(EXCEL_FILE)
    ws_f = wb_f.active
    wb_p = openpyxl.load_workbook(PROSPECT_FILE)
    ws_p = wb_p.active

    # Dates
    now = datetime.now()
    if timeframe == "weekly":
        days = 7
    elif timeframe == "monthly":
        days = 30
    else:
        days = 7

    # Prepare counts per day
    freelancer_counts = []
    prospect_counts = []
    dates = []

    for i in range(days):
        day = now - timedelta(days=days-i-1)
        dates.append(day.strftime("%d-%b"))
        # Count freelancers
        count_f = len([r for r in ws_f.iter_rows(min_row=2, values_only=True) 
                       if r[6] and datetime.strptime(r[6], "%Y-%m-%d %H:%M:%S").date() == day.date()])
        freelancer_counts.append(count_f)
        # Count prospects
        count_p = len([r for r in ws_p.iter_rows(min_row=2, values_only=True) 
                       if r[5] and datetime.strptime(r[5], "%Y-%m-%d %H:%M:%S").date() == day.date()])
        prospect_counts.append(count_p)

    # Plot
    plt.figure(figsize=(6,4))
    plt.plot(dates, freelancer_counts, marker='o', label='Freelancers')
    plt.plot(dates, prospect_counts, marker='o', label='Prospects')
    plt.title(f"{timeframe.capitalize()} Registration Trend")
    plt.xlabel("Date")
    plt.ylabel("Count")
    plt.legend()
    plt.grid(True)
    trend_img = "trend.png"
    plt.tight_layout()
    plt.savefig(trend_img)
    plt.close()

    # Save PDF
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", "B", 16)
    pdf.cell(0, 10, f"{timeframe.capitalize()} Registration Trend", ln=True, align="C")
    pdf.ln(10)
    pdf.image(trend_img, x=10, y=30, w=190)
    pdf.output(file_name)
    os.remove(trend_img)
    return file_name


async def send_admin_reports(app):
    try:
        wb_f = openpyxl.load_workbook(EXCEL_FILE)
        ws_f = wb_f.active
        wb_p = openpyxl.load_workbook(PROSPECT_FILE)
        ws_p = wb_p.active

        # --- Daily new freelancers & prospects ---
        new_freelancers = filter_new_freelancers(ws_f, 1)
        new_prospects = filter_recent_rows(ws_p, 5, 1)  # you already have this

        # --- Save daily Excel files temporarily ---
        daily_f_file = "daily_new_freelancers.xlsx"
        daily_p_file = "daily_new_prospects.xlsx"

        if new_freelancers:
            wb_new_f = openpyxl.Workbook()
            ws_new_f = wb_new_f.active
            ws_new_f.append([cell.value for cell in ws_f[1]])  # headers
            for row in new_freelancers:
                ws_new_f.append(row)
            wb_new_f.save(daily_f_file)

        if new_prospects:
            wb_new_p = openpyxl.Workbook()
            ws_new_p = wb_new_p.active
            ws_new_p.append([cell.value for cell in ws_p[1]])  # headers
            for row in new_prospects:
                ws_new_p.append(row)
            wb_new_p.save(daily_p_file)

        # --- Send files to admin ---
        if new_freelancers:
            await app.bot.send_document(chat_id=ADMIN_ID, document=open(daily_f_file, "rb"), caption="📈 New Freelancers (last 24h)")
            os.remove(daily_f_file)
        if new_prospects:
            await app.bot.send_document(chat_id=ADMIN_ID, document=open(daily_p_file, "rb"), caption="📈 New Prospects (last 24h)")
            os.remove(daily_p_file)

        # --- Send weekly and monthly trend PDFs ---
            weekly_pdf = create_trend_pdf("weekly", "weekly_trend.pdf")
            monthly_pdf = create_trend_pdf("monthly", "monthly_trend.pdf")

            await app.bot.send_document(chat_id=ADMIN_ID, document=open(weekly_pdf, "rb"), caption="📊 Weekly Registration Trend")
            os.remove(weekly_pdf)

            await app.bot.send_document(chat_id=ADMIN_ID, document=open(monthly_pdf, "rb"), caption="📊 Monthly Registration Trend")
            os.remove(monthly_pdf)


            

        # --- Weekly & monthly PDF can be generated separately ---
        # (we'll add this next)
    except Exception as e:
        print(f"Admin report error: {e}")


# --- Start command with button ---
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    telegram_id = update.message.from_user.id

    if telegram_id == ADMIN_ID:
        # Admin menu
        keyboard = [["Download All Freelancers"], ["Download All Prospects"], ["Broadcast Message"]]
        reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        await update.message.reply_text(
            "👋 Welcome Admin! Here you can download full lists or wait for automatic reports.",
            reply_markup=reply_markup
        )
        return ConversationHandler.END
    else:
    # Regular freelancer
        telegram_id = update.message.from_user.id

    # Check if user already registered
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    registered = False
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[5] == telegram_id:
            registered = True
            break
    wb.close()

    if registered:
        keyboard = [["Add Prospect"], ["Download Prospect List"], ["See Profile"]]
        reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        await update.message.reply_text(
            "👋 Welcome back! You are already registered.",
            reply_markup=reply_markup
        )
    else:
        keyboard = [[InlineKeyboardButton("Register", callback_data="register")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await update.message.reply_text(
            "Welcome! Click the button below to register as a freelancer.",
            reply_markup=reply_markup
        )

    # If admin starts the bot, show admin menu
    if update.message.from_user.id == ADMIN_ID:
        keyboard = [["Broadcast Message"], ["All Prospect List"], ["All Freelancer List"]]
        reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        await update.message.reply_text(
            "👋 Welcome, Admin!\nChoose an option below:",
            reply_markup=reply_markup
        )


def assign_to_officer(freelancer_data):
    """
    Assigns a freelancer to an officer Excel file in round-robin.
    Returns the officer file they were assigned to.
    """
    counts = []
    for file in OFFICER_FILES:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        counts.append(ws.max_row - 1)
        wb.close()

    officer_index = counts.index(min(counts))
    officer_file = OFFICER_FILES[officer_index]

    wb = openpyxl.load_workbook(officer_file)
    ws = wb.active
    ws.append([
    freelancer_data[0],  # Name
    freelancer_data[1],  # Phone
    freelancer_data[2],  # City
    freelancer_data[3],  # NID
    freelancer_data[4],  # TIN
    freelancer_data[5]   # Telegram ID
])

    wb.save(officer_file)

    return officer_file  # return the assigned officer file


# --- Button handler ---
async def button_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data == "register":
        await query.message.reply_text("Please enter your full name:")
        return NAME

# --- Step handlers ---
async def get_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['name'] = update.message.text
    await update.message.reply_text("Enter your phone number:")
    return PHONE

async def get_phone(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['phone'] = update.message.text
    await update.message.reply_text("Enter your city:")
    return CITY

async def get_city(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['city'] = update.message.text
    await update.message.reply_text("Enter your National ID:")
    return NID

async def get_nid(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['nid'] = update.message.text
    await update.message.reply_text("Enter your TIN number:")
    return TIN

async def get_tin(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['tin'] = update.message.text

    # --- Save to Excel ---
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    telegram_id = update.message.from_user.id
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[5] == telegram_id:
            await update.message.reply_text("⚠️ You are already registered!")
            # Show menu even if already registered
            keyboard = [["Add Prospect"], ["Download Prospect List"], ["See Profile"]]
            reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
            await update.message.reply_text("Choose an option from the menu:", reply_markup=reply_markup)
            return ConversationHandler.END

    date_registered = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    ws.append([
        context.user_data['name'],
        context.user_data['phone'],
        context.user_data['city'],
        context.user_data['nid'],
        context.user_data['tin'],
        telegram_id,
        date_registered,
        ""  # Verified column
    ])
    wb.save(EXCEL_FILE)

    # --- Assign officer ---
    assigned_officer_file = assign_to_officer([
        context.user_data['name'],
        context.user_data['phone'],
        context.user_data['city'],
        context.user_data['nid'],
        context.user_data['tin'],
        telegram_id
    ])

    await update.message.reply_text("✅ You are successfully registered!")
    officer = OFFICERS.get(assigned_officer_file)
    if officer:
        await update.message.reply_text(
        f"📝 Your assigned officer is:\nName: {officer['name']}\nPhone: {officer['phone']}\nTelegram: {officer['username']}"
    )

    # --- Confirmation message ---
    

    # --- Show freelancer menu ---
    keyboard = [["Add Prospect"], ["Download Prospect List"], ["See Profile"]]
    reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    await update.message.reply_text("Choose an option from the menu:", reply_markup=reply_markup)

    return ConversationHandler.END



async def menu_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text

    if text == "Add Prospect":
        await update.message.reply_text("Enter Prospect Name:")
        return PROSPECT_NAME
    elif text == "Download Prospect List":
        await download_prospect_list(update, context)
        return ConversationHandler.END
    elif text == "See Profile":
        # Call the full profile function
        await see_profile(update, context)
        return ConversationHandler.END
    elif text == "Broadcast Message" and update.message.from_user.id == ADMIN_ID:
        keyboard = [["✅ Verified Freelancers"], ["⚠️ All Freelancers"]]
        reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        await update.message.reply_text("Choose who you want to broadcast to:", reply_markup=reply_markup)
        return BROADCAST_CHOICE

    else:
        return ConversationHandler.END


async def broadcast_choice(update: Update, context: ContextTypes.DEFAULT_TYPE):
    choice = update.message.text
    if choice == "✅ Verified Freelancers":
        context.user_data["broadcast_target"] = "verified"
    else:
        context.user_data["broadcast_target"] = "all"

    await update.message.reply_text("Please send the message you want to broadcast:")
    return BROADCAST_MESSAGE


async def broadcast_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    message_text = update.message.text
    target = context.user_data["broadcast_target"]

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        telegram_id = row[5]
        verified_status = (len(row) > 7 and row[7] == "Verified")

        if target == "verified" and not verified_status:
            continue

        try:
            await update.get_bot().send_message(chat_id=telegram_id, text=message_text)
            count += 1
        except:
            continue

    wb.close()

    await update.message.reply_text(f"✅ Broadcast sent successfully to {count} freelancers.")

    # Return to admin menu
   # Return to admin menu (fixed)
    keyboard = [
        ["Download All Freelancers"], 
        ["Download All Prospects"], 
        ["Broadcast Message"]
]
    reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    await update.message.reply_text("Choose an option:", reply_markup=reply_markup)


    return ConversationHandler.END


async def get_prospect_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['prospect_name'] = update.message.text
    await update.message.reply_text("Enter Prospect Phone Number:")
    return PROSPECT_PHONE

async def get_prospect_phone(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['prospect_phone'] = update.message.text
    keyboard = [["Home"], ["Shop"], ["Share"]]
    reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)
    await update.message.reply_text("Select Prospect Interest:", reply_markup=reply_markup)
    return PROSPECT_INTEREST

async def get_prospect_interest(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['prospect_interest'] = update.message.text
    await update.message.reply_text("Add any comment about the prospect (or type '-' if none):")
    return PROSPECT_COMMENT

async def get_prospect_comment(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['prospect_comment'] = update.message.text

    wb = openpyxl.load_workbook(PROSPECT_FILE)
    ws = wb.active

    prospect_name = context.user_data['prospect_name']
    prospect_phone = context.user_data['prospect_phone']
    now = datetime.now()  # current date and time

    # --- Check for duplicates within 90 days ---
    for row in ws.iter_rows(min_row=2, values_only=True):
        existing_name = row[1]
        existing_phone = row[2]
        date_added_str = row[5]  # The "Date Added" column
        if date_added_str:
            date_added = datetime.strptime(date_added_str, "%Y-%m-%d %H:%M:%S")
            if (existing_name == prospect_name or existing_phone == prospect_phone) and (now - date_added).days < 90:
                await update.message.reply_text("⚠️ This prospect was already added in the last 90 days!")
                # Show menu again
                keyboard = [["Add Prospect"], ["Download Prospect List"], ["See Profile"]]
                reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
                await update.message.reply_text("Choose an option from the menu:", reply_markup=reply_markup)
                return ConversationHandler.END

    # --- Save new prospect with timestamp ---
    ws.append([
        update.message.from_user.id,
        prospect_name,
        prospect_phone,
        context.user_data['prospect_interest'],
        context.user_data['prospect_comment'],
        now.strftime("%Y-%m-%d %H:%M:%S")  # Save date and time
    ])
    wb.save(PROSPECT_FILE)

    await update.message.reply_text("✅ Prospect added successfully!")

    # Show menu again
    keyboard = [["Add Prospect"], ["Download Prospect List"], ["See Profile"]]
    reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    await update.message.reply_text("Choose an option from the menu:", reply_markup=reply_markup)

    return ConversationHandler.END


async def download_prospect_list(update: Update, context: ContextTypes.DEFAULT_TYPE):
    telegram_id = update.message.from_user.id

    wb = openpyxl.load_workbook(PROSPECT_FILE)
    ws = wb.active

    # --- Filter prospects for this freelancer ---
    prospects = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == telegram_id:
            prospects.append({
                "name": row[1],
                "phone": row[2],
                "interest": row[3],
                "comment": row[4]
            })

    if not prospects:
        await update.message.reply_text("You have no prospects yet.")
        return

    # --- Create PDF ---
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", "B", 14)
    pdf.cell(0, 10, "Your Prospect List", ln=True, align="C")
    pdf.ln(10)

    pdf.set_font("Arial", "", 12)
    for i, p in enumerate(prospects, start=1):
        pdf.cell(0, 8, f"{i}. Name: {p['name']}", ln=True)
        pdf.cell(0, 8, f"   Phone: {p['phone']}", ln=True)
        pdf.cell(0, 8, f"   Interest: {p['interest']}", ln=True)
        pdf.cell(0, 8, f"   Comment: {p['comment']}", ln=True)
        pdf.ln(5)

    # --- Save PDF temporarily ---
    pdf_file = f"prospects_{telegram_id}.pdf"
    pdf.output(pdf_file)

    # --- Send PDF to user ---
    await update.message.reply_document(document=open(pdf_file, "rb"))

    # Optional: Delete the PDF after sending
    os.remove(pdf_file)

# --- See Profile handler ---
async def see_profile(update: Update, context: ContextTypes.DEFAULT_TYPE):
    telegram_id = update.message.from_user.id

    # --- Load freelancer data ---
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    freelancer_name = None
    verified_status = "⚠️ Not Verified"
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[5] == telegram_id:  # Telegram ID column
            freelancer_name = row[0]  # Name column
            verified_status = "✅ Verified" if len(row) > 6 and row[6] == "Verified" else "⚠️ Not Verified"
            break
    wb.close()

    if not freelancer_name:
        await update.message.reply_text("⚠️ You are not registered yet. Please register first.")
        return

    # --- Find assigned officer ---
    assigned_officer = None
    for officer_file in OFFICER_FILES:
        wb_off = openpyxl.load_workbook(officer_file)
        ws_off = wb_off.active
        for row_off in ws_off.iter_rows(min_row=2, values_only=True):
            if row_off[5] == telegram_id:  # Telegram ID column
                assigned_officer = OFFICERS.get(officer_file)
                break
        wb_off.close()
        if assigned_officer:
            break

    # --- Build profile message ---
    msg = f"👤 Profile Info:\n\nName: {freelancer_name}\nTelegram ID: {telegram_id}\nStatus: {verified_status}\n"
    
    if assigned_officer:
        msg += (
            f"\n📝 Assigned Officer:\n"
            f"Name: {assigned_officer['name']}\n"
            f"Phone: {assigned_officer['phone']}\n"
            f"Telegram: {assigned_officer['username']}"
        )
    else:
        msg += "\n📝 No officer assigned yet."

    await update.message.reply_text(msg)

    # --- Show menu again ---
    keyboard = [["Add Prospect"], ["Download Prospect List"], ["See Profile"]]
    reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    await update.message.reply_text("Choose an option from the menu:", reply_markup=reply_markup)




# --- Admin download handlers ---
async def admin_download_all(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    if update.message.from_user.id != ADMIN_ID:
        await update.message.reply_text("❌ You are not authorized.")
        return

    await update.message.reply_text("📁 Preparing your file...")

    if "Freelancers" in text:
        await update.message.reply_document(document=open(EXCEL_FILE, "rb"), filename="freelancers.xlsx")
    elif "Prospects" in text:
        await update.message.reply_document(document=open(PROSPECT_FILE, "rb"), filename="prospects.xlsx")




# --- Cancel handler ---
async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Registration canceled.")
    return ConversationHandler.END

# --- Conversation Handler ---
conv_handler = ConversationHandler(
    entry_points=[CallbackQueryHandler(button_handler, pattern="^register$")],
    states={
        NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_name)],
        PHONE: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_phone)],
        CITY: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_city)],
        NID: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_nid)],
        TIN: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_tin)],
    },
    fallbacks=[CommandHandler('cancel', cancel)],
)

prospect_handler = ConversationHandler(
    entry_points=[MessageHandler(filters.TEXT & ~filters.COMMAND, menu_handler)],
    states={
        PROSPECT_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_prospect_name)],
        PROSPECT_PHONE: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_prospect_phone)],
        PROSPECT_INTEREST: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_prospect_interest)],
        PROSPECT_COMMENT: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_prospect_comment)],
    },
    fallbacks=[CommandHandler("cancel", cancel)]
)

broadcast_handler = ConversationHandler(
    entry_points=[MessageHandler(filters.Regex("^Broadcast Message$"), menu_handler)],
    states={
        BROADCAST_CHOICE: [MessageHandler(filters.TEXT & ~filters.COMMAND, broadcast_choice)],
        BROADCAST_MESSAGE: [MessageHandler(filters.TEXT & ~filters.COMMAND, broadcast_message)],
    },
    fallbacks=[CommandHandler("cancel", cancel)],
)


# --- Bot Setup ---
app = ApplicationBuilder().token("8406016067:AAHsUdEVKhf7-yOnq8HDDvmR49papR_ZDIo").build()
app.add_handler(CommandHandler("start", start))
app.add_handler(conv_handler)
app.add_handler(MessageHandler(
    filters.Regex("^(Download All Freelancers|Download All Prospects|All Freelancer List|All Prospect List)$"),
    admin_download_all
))

app.add_handler(prospect_handler)
# ✅ New and improved broadcast handler
broadcast_handler = ConversationHandler(
    entry_points=[MessageHandler(filters.Regex("^Broadcast Message$"), menu_handler)],
    states={
        BROADCAST_CHOICE: [MessageHandler(filters.TEXT & ~filters.COMMAND, broadcast_choice)],
        BROADCAST_MESSAGE: [MessageHandler(filters.TEXT & ~filters.COMMAND, broadcast_message)],
    },
    fallbacks=[CommandHandler("cancel", cancel)],
)

app.add_handler(broadcast_handler)



# --- Schedule Daily Report (6 AM) ---
async def scheduler(app):
    while True:
        now = datetime.now()
        target = now.replace(hour=6, minute=0, second=0, microsecond=0)
        if now > target:
            target += timedelta(days=1)
        wait_seconds = (target - now).total_seconds()
        await asyncio.sleep(wait_seconds)
        await send_admin_reports(app)

app.job_queue.run_once(lambda ctx: asyncio.create_task(scheduler(app)), 1)
app.job_queue.run_once(lambda ctx: asyncio.create_task(weekly_prospect_reminder(app)), 1)


async def verify_freelancer(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.from_user.id != ADMIN_ID:
        await update.message.reply_text("⛔ You are not authorized to verify freelancers.")
        return

    if len(context.args) != 1:
        await update.message.reply_text("Usage: /verify <telegram_id>")
        return

    target_id = int(context.args[0])
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if row[5].value == target_id:
            row[7].value = "Verified"
            wb.save(EXCEL_FILE)

            await update.message.reply_text(f"✅ Freelancer {target_id} has been verified.")

            # Notify freelancer
            try:
                await context.bot.send_message(
                    chat_id=target_id,
                    text="🎉 Congratulations! Your account has been verified by Ayat Office."
                )
            except:
                pass
            return

    await update.message.reply_text("❌ Freelancer not found.")

async def unverify_freelancer(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.from_user.id != ADMIN_ID:
        await update.message.reply_text("⛔ You are not authorized to unverify freelancers.")
        return

    if len(context.args) != 1:
        await update.message.reply_text("Usage: /unverify <telegram_id>")
        return

    target_id = int(context.args[0])
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if row[5].value == target_id:
            row[6].value = ""
            wb.save(EXCEL_FILE)
            await update.message.reply_text(f"⚠️ Freelancer {target_id} has been unverified.")
            return

    await update.message.reply_text("❌ Freelancer not found.")

app.add_handler(CommandHandler("verify", verify_freelancer))
app.add_handler(CommandHandler("unverify", unverify_freelancer))

# --- Weekly Prospect Reminder (Every Monday 8 AM) ---
async def weekly_prospect_reminder(app):
    while True:
        now = datetime.now()
        # Set next Monday 8 AM
        target = now + timedelta(days=(7 - now.weekday()))  # next Monday
        target = target.replace(hour=8, minute=0, second=0, microsecond=0)
        if now > target:
            target += timedelta(days=7)  # next Monday if already past 8 AM
        wait_seconds = (target - now).total_seconds()
        await asyncio.sleep(wait_seconds)

        try:
            wb_f = openpyxl.load_workbook(EXCEL_FILE)
            ws_f = wb_f.active
            wb_p = openpyxl.load_workbook(PROSPECT_FILE)
            ws_p = wb_p.active

            now_date = datetime.now().date()
            one_week_ago = now_date - timedelta(days=7)

            latest_prospect = {}
            for row in ws_p.iter_rows(min_row=2, values_only=True):
                freelancer_id = row[0]
                date_added_str = row[5]
                if date_added_str:
                    try:
                        date_added = datetime.strptime(date_added_str, "%Y-%m-%d %H:%M:%S").date()
                        if freelancer_id not in latest_prospect or date_added > latest_prospect[freelancer_id]:
                            latest_prospect[freelancer_id] = date_added
                    except:
                        continue

            count = 0
            for row in ws_f.iter_rows(min_row=2, values_only=True):
                freelancer_id = row[5]
                last_added = latest_prospect.get(freelancer_id)
                if not last_added or last_added < one_week_ago:
                    try:
                        await app.bot.send_message(
                            chat_id=freelancer_id,
                            text="📢 Hey! Don’t forget to add your new prospects for this week!"
                        )
                        count += 1
                        await asyncio.sleep(0.1)
                    except Exception as e:
                        print(f"Failed to send reminder to {freelancer_id}: {e}")
            print(f"Weekly prospect reminders sent to {count} freelancers.")
        except Exception as e:
            print(f"Error in weekly reminder: {e}")


# --- Run Bot ---
app.run_polling()
