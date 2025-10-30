from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardMarkup
from datetime import datetime, timedelta
from fpdf import FPDF
from telegram.ext import (
    ApplicationBuilder, CommandHandler, MessageHandler,
    ConversationHandler, ContextTypes, filters, CallbackQueryHandler
)
import openpyxl
import os
import asyncio
from telegram.constants import ChatAction


# --- Admin setup ---
ADMIN_ID = 1971668737 # Replace with your admin Telegram ID

# --- Conversation states -
NAME, PHONE, CITY, NID, TIN = range(5)
# For prospect adding
PROSPECT_NAME, PROSPECT_PHONE, PROSPECT_INTEREST, PROSPECT_COMMENT = range(100, 104)
BROADCAST = 200


# --- Excel setup ---
EXCEL_FILE = "freelancers.xlsx"
PROSPECT_FILE = "prospects.xlsx"

if not os.path.exists(EXCEL_FILE):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Full Name", "Phone Number", "City", "National ID", "TIN Number", "Telegram ID", "Verified"])
    wb.save(EXCEL_FILE)


if not os.path.exists(PROSPECT_FILE):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Freelancer Telegram ID", "Prospect Name", "Phone Number", "Interest", "Comment", "Date Added"])
    wb.save(PROSPECT_FILE)

# --- Officer Excel files setup ---
OFFICER_FILES = ["officer1.xlsx", "officer2.xlsx", "officer3.xlsx", "officer4.xlsx"]
for file in OFFICER_FILES:
    if not os.path.exists(file):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Full Name", "Phone Number", "City", "National ID", "TIN Number", "Telegram ID"])
        wb.save(file)

# --- Admin Report Functions ---

def filter_recent_rows(ws, date_col_index, days):
    """Filter rows from Excel within given number of days."""
    now = datetime.now()
    cutoff = now - timedelta(days=days)
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        date_str = row[date_col_index]
        if not date_str:
            continue
        try:
            date = datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S")
            if date >= cutoff:
                rows.append(row)
        except:
            continue
    return rows


def create_summary_pdf(daily_count, weekly_count, monthly_count, file_name="summary.pdf"):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", "B", 16)
    pdf.cell(0, 10, "Freelancer & Prospect Summary Report", ln=True, align="C")
    pdf.ln(10)
    pdf.set_font("Arial", "", 12)
    pdf.cell(0, 10, f"Daily (Last 24 hrs): {daily_count}", ln=True)
    pdf.cell(0, 10, f"Weekly (Last 7 days): {weekly_count}", ln=True)
    pdf.cell(0, 10, f"Monthly (Last 30 days): {monthly_count}", ln=True)
    pdf.output(file_name)
    return file_name


async def send_admin_reports(app):
    """Generate and send daily, weekly, and monthly reports to admin."""
    try:
        # Load Excel files
        wb_f = openpyxl.load_workbook(EXCEL_FILE)
        ws_f = wb_f.active

        wb_p = openpyxl.load_workbook(PROSPECT_FILE)
        ws_p = wb_p.active

        # Counts for freelancers (based on last column: Telegram ID, no date so skip)
        freelancer_count = ws_f.max_row - 1

        # Counts for prospects (based on date column 6)
        daily_prospects = len(filter_recent_rows(ws_p, 5, 1))
        weekly_prospects = len(filter_recent_rows(ws_p, 5, 7))
        monthly_prospects = len(filter_recent_rows(ws_p, 5, 30))

        # Create summary PDF
        pdf_path = create_summary_pdf(daily_prospects, weekly_prospects, monthly_prospects)

        # Send to admin
        await app.bot.send_chat_action(chat_id=ADMIN_ID, action=ChatAction.UPLOAD_DOCUMENT)
        await app.bot.send_document(chat_id=ADMIN_ID, document=open(pdf_path, "rb"), caption="📊 Daily/Weekly/Monthly Summary Report")
        os.remove(pdf_path)
    except Exception as e:
        print(f"Admin report error: {e}")


# --- Start command with button ---
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    telegram_id = update.message.from_user.id

    if telegram_id == ADMIN_ID:
        # Admin menu
        keyboard = [["Download All Freelancers"], ["Download All Prospects"], ["Broadcast Message"]]
        reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        await update.message.reply_text(
            "👋 Welcome Admin! Here you can download full lists or wait for automatic reports.",
            reply_markup=reply_markup
        )
        return ConversationHandler.END
    else:
    # Regular freelancer
        telegram_id = update.message.from_user.id

    # Check if user already registered
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    registered = False
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[5] == telegram_id:
            registered = True
            break
    wb.close()

    if registered:
        keyboard = [["Add Prospect"], ["Download Prospect List"], ["See Profile"]]
        reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        await update.message.reply_text(
            "👋 Welcome back! You are already registered.",
            reply_markup=reply_markup
        )
    else:
        keyboard = [[InlineKeyboardButton("Register", callback_data="register")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await update.message.reply_text(
            "Welcome! Click the button below to register as a freelancer.",
            reply_markup=reply_markup
        )



def assign_to_officer(freelancer_data):
    """
    Assigns a freelancer to an officer Excel file in round-robin.
    freelancer_data = [Full Name, Phone, City, NID, TIN, Telegram ID]
    """
    # Count total freelancers in each officer file
    counts = []
    for file in OFFICER_FILES:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        counts.append(ws.max_row - 1)  # minus header
        wb.close()

    # Find the officer with minimum count (round-robin effect)
    officer_index = counts.index(min(counts))
    officer_file = OFFICER_FILES[officer_index]

    # Append freelancer to that officer's file
    wb = openpyxl.load_workbook(officer_file)
    ws = wb.active
    ws.append(freelancer_data)
    wb.save(officer_file)

# --- Button handler ---
async def button_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data == "register":
        await query.message.reply_text("Please enter your full name:")
        return NAME

# --- Step handlers ---
async def get_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['name'] = update.message.text
    await update.message.reply_text("Enter your phone number:")
    return PHONE

async def get_phone(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['phone'] = update.message.text
    await update.message.reply_text("Enter your city:")
    return CITY

async def get_city(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['city'] = update.message.text
    await update.message.reply_text("Enter your National ID:")
    return NID

async def get_nid(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['nid'] = update.message.text
    await update.message.reply_text("Enter your TIN number:")
    return TIN

async def get_tin(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['tin'] = update.message.text

    # --- Save to main freelancers Excel ---
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    telegram_id = update.message.from_user.id
    # Prevent duplicate Telegram ID entries
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[5] == telegram_id:
            await update.message.reply_text("⚠️ You are already registered!")
            return ConversationHandler.END

    # Save freelancer to main file
    freelancer_data = [
        context.user_data['name'],
        context.user_data['phone'],
        context.user_data['city'],
        context.user_data['nid'],
        context.user_data['tin'],
        telegram_id,
        ""
    ]
    ws.append(freelancer_data)
    wb.save(EXCEL_FILE)

    # --- Assign freelancer to officer in round-robin ---
    assign_to_officer(freelancer_data)

    # --- Registration complete message with menu keyboard ---
    keyboard = [["Add Prospect"], ["Download Prospect List"], ["See Profile"]]
    reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    await update.message.reply_text(
        "✅ You are successfully registered! Choose an option from the menu:",
        reply_markup=reply_markup
    )

    return ConversationHandler.END



async def menu_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text

    if text == "Add Prospect":
        await update.message.reply_text("Enter Prospect Name:")
        return PROSPECT_NAME
    elif text == "Download Prospect List":
        await download_prospect_list(update, context)
        return ConversationHandler.END
    elif text == "See Profile":
        # Inline quick profile (shows Verified status)
        telegram_id = update.message.from_user.id
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[5] == telegram_id:
                name = row[0]
                verified_status = "✅ Verified" if len(row) > 6 and row[6] == "Verified" else "⚠️ Not Verified"
                await update.message.reply_text(
                    f"👤 Name: {name}\n🆔 Telegram ID: {telegram_id}\n{verified_status}"
                )
                return ConversationHandler.END
        await update.message.reply_text("❌ Profile not found. Please register first.")
        return ConversationHandler.END
    else:
        # fallback for other text messages handled by prospect_handler
        return ConversationHandler.END



async def get_prospect_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['prospect_name'] = update.message.text
    await update.message.reply_text("Enter Prospect Phone Number:")
    return PROSPECT_PHONE

async def get_prospect_phone(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['prospect_phone'] = update.message.text
    keyboard = [["Home"], ["Shop"], ["Share"]]
    reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)
    await update.message.reply_text("Select Prospect Interest:", reply_markup=reply_markup)
    return PROSPECT_INTEREST

async def get_prospect_interest(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['prospect_interest'] = update.message.text
    await update.message.reply_text("Add any comment about the prospect (or type '-' if none):")
    return PROSPECT_COMMENT

async def get_prospect_comment(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['prospect_comment'] = update.message.text

    wb = openpyxl.load_workbook(PROSPECT_FILE)
    ws = wb.active

    prospect_name = context.user_data['prospect_name']
    prospect_phone = context.user_data['prospect_phone']
    now = datetime.now()  # current date and time

    # --- Check for duplicates within 90 days ---
    for row in ws.iter_rows(min_row=2, values_only=True):
        existing_name = row[1]
        existing_phone = row[2]
        date_added_str = row[5]  # The "Date Added" column
        if date_added_str:
            date_added = datetime.strptime(date_added_str, "%Y-%m-%d %H:%M:%S")
            if (existing_name == prospect_name or existing_phone == prospect_phone) and (now - date_added).days < 90:
                await update.message.reply_text("⚠️ This prospect was already added in the last 90 days!")
                # Show menu again
                keyboard = [["Add Prospect"], ["Download Prospect List"], ["See Profile"]]
                reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
                await update.message.reply_text("Choose an option from the menu:", reply_markup=reply_markup)
                return ConversationHandler.END

    # --- Save new prospect with timestamp ---
    ws.append([
        update.message.from_user.id,
        prospect_name,
        prospect_phone,
        context.user_data['prospect_interest'],
        context.user_data['prospect_comment'],
        now.strftime("%Y-%m-%d %H:%M:%S")  # Save date and time
    ])
    wb.save(PROSPECT_FILE)

    await update.message.reply_text("✅ Prospect added successfully!")

    # Show menu again
    keyboard = [["Add Prospect"], ["Download Prospect List"], ["See Profile"]]
    reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    await update.message.reply_text("Choose an option from the menu:", reply_markup=reply_markup)

    return ConversationHandler.END


async def download_prospect_list(update: Update, context: ContextTypes.DEFAULT_TYPE):
    telegram_id = update.message.from_user.id

    wb = openpyxl.load_workbook(PROSPECT_FILE)
    ws = wb.active

    # --- Filter prospects for this freelancer ---
    prospects = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == telegram_id:
            prospects.append({
                "name": row[1],
                "phone": row[2],
                "interest": row[3],
                "comment": row[4]
            })

    if not prospects:
        await update.message.reply_text("You have no prospects yet.")
        return

    # --- Create PDF ---
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", "B", 14)
    pdf.cell(0, 10, "Your Prospect List", ln=True, align="C")
    pdf.ln(10)

    pdf.set_font("Arial", "", 12)
    for i, p in enumerate(prospects, start=1):
        pdf.cell(0, 8, f"{i}. Name: {p['name']}", ln=True)
        pdf.cell(0, 8, f"   Phone: {p['phone']}", ln=True)
        pdf.cell(0, 8, f"   Interest: {p['interest']}", ln=True)
        pdf.cell(0, 8, f"   Comment: {p['comment']}", ln=True)
        pdf.ln(5)

    # --- Save PDF temporarily ---
    pdf_file = f"prospects_{telegram_id}.pdf"
    pdf.output(pdf_file)

    # --- Send PDF to user ---
    await update.message.reply_document(document=open(pdf_file, "rb"))

    # Optional: Delete the PDF after sending
    os.remove(pdf_file)

# --- See Profile handler ---
async def see_profile(update: Update, context: ContextTypes.DEFAULT_TYPE):
    telegram_id = update.message.from_user.id

    # Load freelancers Excel
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    # Find freelancer info
    name = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[5] == telegram_id:
            name = row[0]
            break

    if name:
        await update.message.reply_text(f"👤 Profile Info:\n\nName: {name}\nTelegram ID: {telegram_id}")
    else:
        await update.message.reply_text("⚠️ You are not registered yet. Please register first.")

    # Show menu again
    keyboard = [["Add Prospect"], ["Download Prospect List"], ["See Profile"]]
    reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    await update.message.reply_text("Choose an option from the menu:", reply_markup=reply_markup)

# --- Admin download handlers ---
async def admin_download_all(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    if update.message.from_user.id != ADMIN_ID:
        await update.message.reply_text("❌ You are not authorized.")
        return

    await update.message.reply_text("📁 Preparing your file...")

    if "Freelancers" in text:
        await update.message.reply_document(document=open(EXCEL_FILE, "rb"), filename="freelancers.xlsx")
    elif "Prospects" in text:
        await update.message.reply_document(document=open(PROSPECT_FILE, "rb"), filename="prospects.xlsx")

# --- Admin broadcast handler ---
async def admin_broadcast_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.from_user.id != ADMIN_ID:
        await update.message.reply_text("❌ You are not authorized.")
        return ConversationHandler.END

    await update.message.reply_text("📢 Please send the message you want to broadcast to all freelancers:")
    return BROADCAST


async def admin_broadcast_send(update: Update, context: ContextTypes.DEFAULT_TYPE):
    message_text = update.message.text
    count = 0

    # Load freelancers
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    # Send to all freelancer Telegram IDs (column 6)
    for row in ws.iter_rows(min_row=2, values_only=True):
        telegram_id = row[5]
        if telegram_id:
            try:
                await update.get_bot().send_message(chat_id=telegram_id, text=message_text)
                count += 1
                await asyncio.sleep(0.1)  # small delay to avoid flood limit
            except Exception as e:
                print(f"Failed to send to {telegram_id}: {e}")

    await update.message.reply_text(f"✅ Broadcast complete! Sent to {count} freelancers.")
    return ConversationHandler.END


# --- Cancel handler ---
async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Registration canceled.")
    return ConversationHandler.END

# --- Conversation Handler ---
conv_handler = ConversationHandler(
    entry_points=[CallbackQueryHandler(button_handler, pattern="^register$")],
    states={
        NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_name)],
        PHONE: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_phone)],
        CITY: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_city)],
        NID: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_nid)],
        TIN: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_tin)],
    },
    fallbacks=[CommandHandler('cancel', cancel)],
)

prospect_handler = ConversationHandler(
    entry_points=[MessageHandler(filters.TEXT & ~filters.COMMAND, menu_handler)],
    states={
        PROSPECT_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_prospect_name)],
        PROSPECT_PHONE: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_prospect_phone)],
        PROSPECT_INTEREST: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_prospect_interest)],
        PROSPECT_COMMENT: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_prospect_comment)],
    },
    fallbacks=[CommandHandler("cancel", cancel)]
)

broadcast_handler = ConversationHandler(
    entry_points=[MessageHandler(filters.Regex("^Broadcast Message$"), admin_broadcast_start)],
    states={
        BROADCAST: [MessageHandler(filters.TEXT & ~filters.COMMAND, admin_broadcast_send)],
    },
    fallbacks=[CommandHandler("cancel", cancel)],
)

# --- Bot Setup ---
app = ApplicationBuilder().token("8406016067:AAHsUdEVKhf7-yOnq8HDDvmR49papR_ZDIo").build()
app.add_handler(CommandHandler("start", start))
app.add_handler(conv_handler)
app.add_handler(MessageHandler(filters.Regex("^(Download All Freelancers|Download All Prospects)$"), admin_download_all))
app.add_handler(broadcast_handler)
app.add_handler(prospect_handler)


# --- Schedule Daily Report (6 AM) ---
async def scheduler(app):
    while True:
        now = datetime.now()
        target = now.replace(hour=6, minute=0, second=0, microsecond=0)
        if now > target:
            target += timedelta(days=1)
        wait_seconds = (target - now).total_seconds()
        await asyncio.sleep(wait_seconds)
        await send_admin_reports(app)

app.job_queue.run_once(lambda ctx: asyncio.create_task(scheduler(app)), 1)

async def verify_freelancer(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.from_user.id != ADMIN_ID:
        await update.message.reply_text("⛔ You are not authorized to verify freelancers.")
        return

    if len(context.args) != 1:
        await update.message.reply_text("Usage: /verify <telegram_id>")
        return

    target_id = int(context.args[0])
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if row[5].value == target_id:
            row[6].value = "Verified"
            wb.save(EXCEL_FILE)

            await update.message.reply_text(f"✅ Freelancer {target_id} has been verified.")

            # Notify freelancer
            try:
                await context.bot.send_message(
                    chat_id=target_id,
                    text="🎉 Congratulations! Your account has been verified by Ayat Office."
                )
            except:
                pass
            return

    await update.message.reply_text("❌ Freelancer not found.")

async def unverify_freelancer(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.from_user.id != ADMIN_ID:
        await update.message.reply_text("⛔ You are not authorized to unverify freelancers.")
        return

    if len(context.args) != 1:
        await update.message.reply_text("Usage: /unverify <telegram_id>")
        return

    target_id = int(context.args[0])
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if row[5].value == target_id:
            row[6].value = ""
            wb.save(EXCEL_FILE)
            await update.message.reply_text(f"⚠️ Freelancer {target_id} has been unverified.")
            return

    await update.message.reply_text("❌ Freelancer not found.")

app.add_handler(CommandHandler("verify", verify_freelancer))
app.add_handler(CommandHandler("unverify", unverify_freelancer))

# --- Run Bot ---
app.run_polling()
